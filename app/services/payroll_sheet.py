"""The four-weekly master sheet, in the payroll workbook's own layout.

This reproduces ``Format.xlsx`` - the sheet the office already sends to payroll -
rather than inventing a new one, so nothing downstream has to change. Columns,
widths, banding, borders and the yellow/amber input shading all match.

What is filled in and what is left alone follows one rule: **anything the
clocking system knows is written; anything it cannot know is left blank for a
human, and anything derivable from another column is a formula rather than a
number.**

    From the clock       Forename, Surname, Department, Payroll ref,
                         Basic hours (J), O/T hours (K), and any timesheet
                         warning in Notes.
    From the database    Basic rate (F), when one has been recorded on the
                         employee's card. Blank otherwise - a wage is never
                         guessed.
    Typed in             Holiday hours (M), SSP days (Q), Back pay (W),
                         Adjustments (X), Deductions (Y), start and leaving
                         dates.
    Excel works out      The O/T 1.50 rate (H), Total hours (O), Basic pay (S),
                         O/T pay (U), Holiday pay (V) and TOTAL PAY (Z).

Because the right-hand columns are live formulas, correcting a rate or typing a
holiday figure re-totals the sheet in front of whoever is checking it; nobody
retypes a number that Excel can work out.

The one rate on the sheet is the basic one, in column F. The "O/T 1.50" rate in
column H is ``=F*1.5``, not a second figure to keep in step - so a rise typed
into F carries straight through to overtime, and the two can never disagree.
That formula is written on every row, including rows where no basic rate is
recorded yet: it is the rule rather than a value, so typing a rate into F in
Excel produces the overtime rate without anybody touching column H.

SSP is deliberately outside TOTAL PAY: the sheet records the number of days, not
a rate, and statutory sick pay is settled by the payroll bureau.

Only four-weekly staff are on the sheet. Salaried staff are paid a fixed amount
whatever they clock, so hours and rates for them would be a wage this system has
no business working out - see :func:`sheet_employees`.
"""

from __future__ import annotations

import datetime as dt
import io
from dataclasses import dataclass

from sqlalchemy import select

from ..extensions import db
from ..models import PAY_FOUR_WEEKLY, PAY_SALARY, Employee, visible_employee_clause
from .payrates import OVERTIME_MULTIPLIER
from .timesheet import EmployeeTotal

# --- the layout, one entry per column -------------------------------------
# Kept as data so the sheet's shape is readable in one place and a change to
# the payroll workbook is a change here rather than a hunt through the code.
FIRST_DATA_ROW = 10
HEADER_TOP_ROW = 7
HEADER_LABEL_ROW = 8
HEADER_PAD_ROW = 9

ACCOUNTING = '_-* #,##0.00_-;\\-* #,##0.00_-;_-* "-"??_-;_-@_-'
ACCOUNTING_WHOLE = '_-* #,##0_-;\\-* #,##0_-;_-* "-"??_-;_-@_-'
UK_DATE = "dd/mm/yy"

GREY = "FFD9D9D9"  # header band
YELLOW = "FFFFFF00"  # rates: confidential, typed in by payroll
AMBER = "FFFFC000"  # hours to pay, and the headline total

MEDIUM = "medium"
THIN = "thin"


@dataclass(frozen=True)
class Column:
    """One column of the payroll sheet."""

    letter: str
    heading: str = ""
    width: float | None = None
    fill: str | None = None  # data-cell shading
    number_format: str | None = None
    align: str | None = None
    left: str | None = None
    right: str | None = None
    bottom: str | None = None  # border closing the table under the last row


# Spacer columns (E, I, N, P, R, AA) carry no border and no heading; they are
# the gaps that separate the banded groups in the original.
COLUMNS: tuple[Column, ...] = (
    Column("A", "Forename", 15.43, left=MEDIUM, right=THIN, bottom=MEDIUM),
    Column("B", "Surname", 15.71, right=THIN, bottom=MEDIUM),
    Column("C", "Department", 13.29, left=THIN, right=THIN, bottom=MEDIUM),
    Column("D", "Payroll Ref", 8.71, number_format=ACCOUNTING_WHOLE, align="right",
           left=THIN, right=MEDIUM, bottom=MEDIUM),
    Column("E"),
    Column("F", "Basic", 4.71, fill=YELLOW, number_format=ACCOUNTING,
           left=MEDIUM, right=THIN, bottom=MEDIUM),
    Column("G", "", None, fill=YELLOW, number_format=ACCOUNTING,
           left=THIN, right=THIN, bottom=MEDIUM),
    Column("H", "O/T 1.50", 7.14, number_format=ACCOUNTING,
           right=MEDIUM, bottom=MEDIUM),
    Column("I"),
    Column("J", "Basic", 4.71, fill=AMBER, number_format=ACCOUNTING, align="center",
           left=MEDIUM, right=THIN, bottom=MEDIUM),
    Column("K", "O/T 1.50", 7.14, fill=AMBER, number_format=ACCOUNTING, align="center",
           left=THIN, right=THIN, bottom=MEDIUM),
    Column("L", "", None, fill=AMBER, number_format=ACCOUNTING, align="center",
           left=THIN, right=THIN, bottom=MEDIUM),
    Column("M", "Holiday", 6.43, fill=AMBER, number_format=ACCOUNTING,
           right=MEDIUM, bottom=MEDIUM),
    Column("N"),
    Column("O", "Total Hours", None, number_format=ACCOUNTING,
           left=MEDIUM, right=MEDIUM, bottom=MEDIUM),
    Column("P"),
    Column("Q", "SSP Days", 7.57, number_format=ACCOUNTING,
           left=MEDIUM, right=MEDIUM, bottom=MEDIUM),
    Column("R"),
    Column("S", "Basic", 4.71, number_format=ACCOUNTING,
           left=MEDIUM, right=THIN, bottom=THIN),
    Column("T", "", None, number_format=ACCOUNTING, right=THIN, bottom=THIN),
    Column("U", "O/T 1.50", 7.14, number_format=ACCOUNTING,
           left=THIN, right=THIN, bottom=THIN),
    Column("V", "Holiday", 6.43, number_format=ACCOUNTING,
           left=THIN, right=THIN, bottom=THIN),
    Column("W", "Back Pay", 7.43, number_format=ACCOUNTING,
           left=THIN, right=THIN, bottom=THIN),
    Column("X", "Adjustments", None, number_format=ACCOUNTING,
           left=THIN, right=THIN, bottom=THIN),
    Column("Y", "Deductions", 8.86, number_format=ACCOUNTING,
           left=THIN, right=MEDIUM, bottom=THIN),
    Column("Z", "TOTAL PAY", 8.71, fill=AMBER, number_format=ACCOUNTING,
           right=MEDIUM, bottom=THIN),
    Column("AA"),
    Column("AB", "Start Date for new staff", 8.29, number_format=UK_DATE, align="center",
           left=MEDIUM, right=THIN, bottom=MEDIUM),
    Column("AC", "Leaving date for leavers", 6.57, number_format=UK_DATE, align="center",
           bottom=MEDIUM),
    Column("AD", "Notes", 5.14, left=THIN, right=MEDIUM, bottom=THIN),
)

BY_LETTER = {column.letter: column for column in COLUMNS}

# The banded group headings above the column headings.
GROUPS: tuple[tuple[str, str, str], ...] = (
    ("Rates of Pay", "F", "H"),
    ("Hours to Pay", "J", "M"),
    ("Total Pay", "S", "Z"),
    ("Notes", "AB", "AD"),
)

PERIOD_DAYS = 28


def payroll_period(start: dt.date, anchor: dt.date, periods_per_year: int = 13) -> int:
    """Which four-weekly period *start* falls in, counting from *anchor*.

    Anchored on a configured date rather than derived from the calendar, because
    a four-weekly payroll year has no fixed relationship to January: only the
    company knows when its period 1 began.
    """
    elapsed = (start - anchor).days
    return (elapsed // PERIOD_DAYS) % periods_per_year + 1


def _short(day: dt.date) -> str:
    return day.strftime("%d/%m/%y")


def period_label(start: dt.date, end: dt.date) -> str:
    """The "17/08/26 - 13/09/26" line under Period."""
    return f"{_short(start)} - {_short(end)}"


def sheet_employees(
    *, employee_id: int | None = None, department: str | None = None
) -> list[Employee]:
    """Everyone who belongs on the sheet, in the order the office keeps them.

    Every active **four-weekly** employee appears, including anyone who clocked
    nothing in the period: payroll still has to put their holiday or sick days
    somewhere, and a name silently missing from a wage sheet is how somebody
    ends up unpaid.

    Salaried staff are the one deliberate exception. They are paid a fixed
    amount whatever they clock, so a row of hours and rates for them would be a
    wage this system has no business working out. They are counted by
    :func:`excluded_salaried` so the exclusion is stated on screen rather than
    being a silent gap in the list.

    Ordered by forename to match the workbook, which is maintained that way.
    """
    stmt = (
        select(Employee)
        .where(
            Employee.is_active.is_(True),
            Employee.pay_basis == PAY_FOUR_WEEKLY,
            visible_employee_clause(),
        )
        .order_by(Employee.first_name, Employee.last_name)
    )
    if employee_id is not None:
        stmt = stmt.where(Employee.id == employee_id)
    if department:
        stmt = stmt.where(Employee.department == department)
    return list(db.session.scalars(stmt).all())


def excluded_salaried(
    *, employee_id: int | None = None, department: str | None = None
) -> list[Employee]:
    """Active salaried staff left off the sheet, for the same filters.

    Exists so the omission can be shown to whoever downloads the sheet. An
    exclusion nobody is told about is indistinguishable from a bug.
    """
    stmt = (
        select(Employee)
        .where(
            Employee.is_active.is_(True),
            Employee.pay_basis == PAY_SALARY,
            visible_employee_clause(),
        )
        .order_by(Employee.first_name, Employee.last_name)
    )
    if employee_id is not None:
        stmt = stmt.where(Employee.id == employee_id)
    if department:
        stmt = stmt.where(Employee.department == department)
    return list(db.session.scalars(stmt).all())


def build_master_workbook(
    totals: list[EmployeeTotal],
    start: dt.date,
    end: dt.date,
    *,
    employees: list[Employee] | None = None,
    period: int | None = None,
    company_name: str = "Moduflex Ltd",
    anchor: dt.date | None = None,
    periods_per_year: int = 13,
):
    """Build the master sheet workbook for one payroll period."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import column_index_from_string

    if employees is None:
        employees = [total.employee for total in totals]
    if period is None and anchor is not None:
        period = payroll_period(start, anchor, periods_per_year)

    by_employee = {total.employee.id: total for total in totals}

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"

    base = Font(name="Calibri", size=9)
    bold = Font(name="Calibri", size=9, bold=True)
    title = Font(name="Calibri", size=9, bold=True, underline="single")
    grey = PatternFill("solid", start_color=GREY, end_color=GREY)
    yellow = PatternFill("solid", start_color=YELLOW, end_color=YELLOW)
    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def side(style: str | None) -> Side:
        return Side(style=style) if style else Side()

    def fill_for(name: str | None) -> PatternFill | None:
        if name is None:
            return None
        return PatternFill("solid", start_color=name, end_color=name)

    # --- title block ------------------------------------------------------
    sheet["A1"] = company_name
    sheet["A1"].font = title
    sheet["A2"] = "Four Weekly Payroll"
    sheet["A2"].font = title

    sheet["A4"] = "Period:"
    sheet["A4"].font = bold
    if period is not None:
        sheet["B4"] = period
    sheet["B4"].font = bold
    sheet["B4"].fill = yellow
    sheet["B4"].alignment = Alignment(horizontal="left")

    sheet["A5"] = "Dates:"
    sheet["A5"].font = bold
    sheet["A5"].alignment = Alignment(horizontal="left")
    sheet["B5"] = period_label(start, end)
    sheet["B5"].font = bold
    sheet["B5"].fill = yellow
    sheet["B5"].alignment = Alignment(horizontal="left")

    # Present in the original workbook and kept for fidelity: an empty merge
    # above the rates band, left over from how the sheet was laid out by hand.
    sheet.merge_cells("F4:H4")

    # --- group band and column headings -----------------------------------
    for heading, first, last in GROUPS:
        sheet.merge_cells(f"{first}{HEADER_TOP_ROW}:{last}{HEADER_TOP_ROW}")
        for index in range(
            column_index_from_string(first), column_index_from_string(last) + 1
        ):
            cell = sheet.cell(row=HEADER_TOP_ROW, column=index)
            cell.fill = grey
            cell.font = bold
            cell.alignment = centre
            cell.border = Border(
                top=side(MEDIUM),
                bottom=side(MEDIUM),
                left=side(MEDIUM if index == column_index_from_string(first) else None),
                right=side(MEDIUM if index == column_index_from_string(last) else None),
            )
        sheet[f"{first}{HEADER_TOP_ROW}"] = heading

    for column in COLUMNS:
        index = column_index_from_string(column.letter)
        if column.width:
            sheet.column_dimensions[column.letter].width = column.width
        if column.left is None and column.right is None and not column.heading:
            continue  # a spacer column: no heading, no band, no border

        for row, text in ((HEADER_LABEL_ROW, column.heading), (HEADER_PAD_ROW, None)):
            cell = sheet.cell(row=row, column=index)
            if text:
                cell.value = text
            cell.font = bold
            cell.fill = grey
            cell.alignment = centre
            cell.border = Border(
                left=side(column.left),
                right=side(column.right),
                top=side(MEDIUM) if row == HEADER_LABEL_ROW else side(None),
            )

    sheet.row_dimensions[HEADER_TOP_ROW].height = 15.75
    sheet.row_dimensions[HEADER_LABEL_ROW].height = 37.5
    sheet.row_dimensions[HEADER_PAD_ROW].height = 15.75

    # --- one row per employee ---------------------------------------------
    last_row = FIRST_DATA_ROW + len(employees) - 1
    for offset, employee in enumerate(employees):
        row = FIRST_DATA_ROW + offset
        total = by_employee.get(employee.id)
        is_last = row == last_row

        for column in COLUMNS:
            if column.left is None and column.right is None and not column.heading:
                continue
            cell = sheet.cell(row=row, column=column_index_from_string(column.letter))
            cell.font = base
            if column.number_format:
                cell.number_format = column.number_format
            if column.fill:
                cell.fill = fill_for(column.fill)
            if column.align:
                cell.alignment = Alignment(horizontal=column.align)
            cell.border = Border(
                left=side(column.left),
                right=side(column.right),
                bottom=side(column.bottom) if is_last else side(None),
            )

        sheet[f"A{row}"] = employee.first_name
        sheet[f"B{row}"] = employee.last_name
        sheet[f"C{row}"] = employee.department or ""
        # A numeric payroll ref sorts and formats as a number; one with letters
        # in it stays text rather than becoming an Excel error.
        reference = employee.payroll_ref
        sheet[f"D{row}"] = int(reference) if reference.isdigit() else reference

        # The basic rate, only where one has actually been recorded. A blank
        # here is a prompt to type the rate in, not a zero to be paid.
        basic_rate = employee.basic_rate
        if basic_rate is not None:
            sheet[f"F{row}"] = float(basic_rate)
        # Overtime is always time and a half on whatever ends up in F, so this
        # is written whether or not a rate is on file yet.
        sheet[f"H{row}"] = f"=F{row}*{OVERTIME_MULTIPLIER:g}"

        # Hours to pay, from the clock. Somebody with no shifts still gets a
        # row, with zeroes, so holiday or sick days can be entered against it.
        sheet[f"J{row}"] = round(total.standard_hours, 2) if total else 0.0
        sheet[f"K{row}"] = round(total.overtime_hours, 2) if total else 0.0

        # Everything to the right reads from the columns to its left.
        sheet[f"O{row}"] = f"=J{row}+K{row}+M{row}"
        sheet[f"S{row}"] = f"=J{row}*F{row}"
        sheet[f"U{row}"] = f"=K{row}*H{row}"
        sheet[f"V{row}"] = f"=M{row}*F{row}"
        sheet[f"Z{row}"] = f"=S{row}+U{row}+V{row}+W{row}+X{row}-Y{row}"

        # Only written when there is something wrong, so the column stays free
        # for whoever is checking the sheet to write on.
        if total and total.issue_details:
            sheet[f"AD{row}"] = "; ".join(total.issue_details)

    return workbook


def to_master_xlsx(
    totals: list[EmployeeTotal],
    start: dt.date,
    end: dt.date,
    **kwargs,
) -> bytes:
    """The master sheet as .xlsx bytes, ready to send back as a download."""
    workbook = build_master_workbook(totals, start, end, **kwargs)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
